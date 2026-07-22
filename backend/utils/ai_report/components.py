"""
Reusable, screenshot-accurate rendering primitives.

Every visual element in the workbook is built from these functions so the seven
sheets stay consistent by construction: title bands, green section headers,
light-blue KPI cards, the navy "callout" line, and the banded data table with an
Excel ``ListObject`` (auto-filter + sortable) on top.

Sheet modules describe *what* to render (a title, some cards, a list of columns
and rows); these helpers decide *how* it looks.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Callable, Optional, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from . import theme as T


# --------------------------------------------------------------------------- #
# Column specification for data tables
# --------------------------------------------------------------------------- #
@dataclass
class Col:
    """One column in a styled data table."""

    header: str
    width: float = 16.0
    align: str = "left"                 # left | center | right
    fmt: Optional[str] = None           # openpyxl number_format
    # value(row) -> cell value. Rows are plain dicts from the dataset layer.
    get: Optional[Callable[[Any], Any]] = None
    key: Optional[str] = None           # convenience: read row[key]

    def value(self, row: Any) -> Any:
        if self.get is not None:
            return self.get(row)
        if self.key is not None:
            if isinstance(row, dict):
                return row.get(self.key)
            return getattr(row, self.key, None)
        return None


_ALIGN = {
    "left": T.ALIGN_LEFT,
    "center": T.ALIGN_CENTER,
    "right": T.ALIGN_RIGHT,
}


# --------------------------------------------------------------------------- #
# Low-level helpers
# --------------------------------------------------------------------------- #
def col_letter(idx: int) -> str:
    return get_column_letter(idx)


def merge(ws: Worksheet, r1: int, c1: int, r2: int, c2: int) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def fill_range(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, fill) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = fill


def set_widths(ws: Worksheet, widths: Sequence[float], start_col: int = 1) -> None:
    for i, w in enumerate(widths):
        ws.column_dimensions[col_letter(start_col + i)].width = w


# --------------------------------------------------------------------------- #
# Title band  (navy full-width banner + italic subtitle)
# --------------------------------------------------------------------------- #
def title_band(
    ws: Worksheet,
    title: str,
    subtitle: str = "",
    *,
    last_col: int,
    row: int = 1,
    first_col: int = 1,
) -> int:
    """Render the navy title banner. Returns the next free row."""
    merge(ws, row, first_col, row, last_col)
    fill_range(ws, row, first_col, row, last_col, T.FILL_TITLE)
    cell = ws.cell(row=row, column=first_col, value=title)
    cell.font = T.FONT_TITLE
    cell.alignment = T.ALIGN_TITLE
    ws.row_dimensions[row].height = 30
    nxt = row + 1
    if subtitle:
        merge(ws, nxt, first_col, nxt, last_col)
        sc = ws.cell(row=nxt, column=first_col, value=subtitle)
        sc.font = T.FONT_SUBTITLE
        sc.alignment = T.ALIGN_TITLE
        ws.row_dimensions[nxt].height = 16
        nxt += 1
    return nxt + 1  # one blank spacer row


# --------------------------------------------------------------------------- #
# Callout line  (e.g. "Overall AI adoption: 72% ...")
# --------------------------------------------------------------------------- #
def callout(ws: Worksheet, text: str, *, row: int, first_col: int = 1, last_col: int) -> int:
    merge(ws, row, first_col, row, last_col)
    cell = ws.cell(row=row, column=first_col, value=text)
    cell.font = T.FONT_CALLOUT
    cell.alignment = T.ALIGN_LEFT
    ws.row_dimensions[row].height = 20
    return row + 2


# --------------------------------------------------------------------------- #
# Section header  (green bar)
# --------------------------------------------------------------------------- #
def section_header(ws: Worksheet, text: str, *, row: int, first_col: int = 1, last_col: int) -> int:
    merge(ws, row, first_col, row, last_col)
    fill_range(ws, row, first_col, row, last_col, T.FILL_SECTION)
    cell = ws.cell(row=row, column=first_col, value=text)
    cell.font = T.FONT_SECTION
    cell.alignment = T.ALIGN_LEFT
    ws.row_dimensions[row].height = 22
    return row + 1


# --------------------------------------------------------------------------- #
# KPI cards  (continuous light-blue band: label row over big-number row)
# --------------------------------------------------------------------------- #
@dataclass
class Kpi:
    label: str
    value: Any
    fmt: Optional[str] = None
    sub: Optional[str] = None


def kpi_cards(
    ws: Worksheet,
    cards: Sequence[Kpi],
    *,
    row: int,
    first_col: int = 1,
    span: int = 2,
) -> int:
    """Lay out KPI cards as one continuous band. Returns the next free row."""
    label_row = row
    value_row = row + 1
    n = len(cards)
    last_col = first_col + span * n - 1

    fill_range(ws, label_row, first_col, value_row, last_col, T.FILL_KPI)
    ws.row_dimensions[label_row].height = 22
    ws.row_dimensions[value_row].height = 34

    for i, card in enumerate(cards):
        c1 = first_col + i * span
        c2 = c1 + span - 1
        merge(ws, label_row, c1, label_row, c2)
        merge(ws, value_row, c1, value_row, c2)

        lc = ws.cell(row=label_row, column=c1, value=card.label)
        lc.font = T.FONT_KPI_LABEL
        lc.alignment = T.ALIGN_KPI_LABEL

        vc = ws.cell(row=value_row, column=c1, value=card.value)
        vc.font = T.FONT_KPI_VALUE
        vc.alignment = T.ALIGN_KPI_VALUE
        if card.fmt and isinstance(card.value, (int, float)):
            vc.number_format = card.fmt

    return value_row + 2  # trailing spacer row


# --------------------------------------------------------------------------- #
# Data table  (navy header + banded body + Excel ListObject)
# --------------------------------------------------------------------------- #
def _unique_headers(cols: Sequence[Col]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for c in cols:
        h = c.header
        if h in seen:
            seen[h] += 1
            h = f"{h} ({seen[c.header]})"
        else:
            seen[c.header] = 0
        out.append(h)
    return out


def data_table(
    ws: Worksheet,
    cols: Sequence[Col],
    rows: Sequence[Any],
    *,
    start_row: int,
    start_col: int = 1,
    table_name: Optional[str] = None,
    row_fill: Optional[Callable[[int, Any], Any]] = None,
    zebra: bool = True,
) -> int:
    """
    Render a fully styled table and (optionally) wrap it in an Excel table.

    ``row_fill(index, row) -> PatternFill|None`` lets a caller tint specific rows
    (used for provider colour-coding and adoption status). It takes precedence
    over zebra striping. Returns the next free row after the table.
    """
    headers = _unique_headers(cols)
    ncol = len(cols)
    header_row = start_row
    end_col = start_col + ncol - 1

    # Header
    for j, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=start_col + j, value=h)
        cell.font = T.FONT_HEADER
        cell.fill = T.FILL_HEADER
        cell.alignment = T.ALIGN_CENTER
        cell.border = T.BORDER_HEADER
    ws.row_dimensions[header_row].height = 28

    # Body
    r = header_row
    for i, row in enumerate(rows):
        r = header_row + 1 + i
        explicit = row_fill(i, row) if row_fill else None
        base = explicit or (T.FILL_ALT if (zebra and i % 2 == 1) else T.FILL_WHITE)
        for j, col in enumerate(cols):
            cell = ws.cell(row=r, column=start_col + j, value=col.value(row))
            cell.font = T.FONT_BODY
            cell.fill = base
            cell.alignment = _ALIGN.get(col.align, T.ALIGN_LEFT)
            cell.border = T.BORDER_CELL
            if col.fmt:
                cell.number_format = col.fmt
    last_row = header_row + len(rows)

    # Column widths
    for j, col in enumerate(cols):
        ws.column_dimensions[col_letter(start_col + j)].width = col.width

    # Excel table (auto-filter + sortable). Requires at least the header row.
    if table_name:
        ref = f"{col_letter(start_col)}{header_row}:{col_letter(end_col)}{max(last_row, header_row + 1)}"
        # If there are zero data rows, extend by one to satisfy Excel's ref rules.
        table = Table(displayName=_safe_table_name(ws, table_name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=T.TABLE_STYLE,
            showRowStripes=False,      # we paint our own zebra / tints
            showColumnStripes=False,
            showFirstColumn=False,
            showLastColumn=False,
        )
        ws.add_table(table)
    else:
        # No ListObject -> still give an auto-filter for sortability.
        ws.auto_filter.ref = (
            f"{col_letter(start_col)}{header_row}:{col_letter(end_col)}{max(last_row, header_row)}"
        )

    return last_row + 2


_used_table_names: set[str] = set()


def _safe_table_name(ws: Worksheet, name: str) -> str:
    """Excel table names must be unique, start with a letter, no spaces."""
    base = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in name)
    if not base or not base[0].isalpha():
        base = "T_" + base
    candidate = base
    i = 1
    key = f"{ws.title}:{candidate}"
    while key in _used_table_names:
        i += 1
        candidate = f"{base}_{i}"
        key = f"{ws.title}:{candidate}"
    _used_table_names.add(key)
    return candidate


# --------------------------------------------------------------------------- #
# Sheet-level conveniences
# --------------------------------------------------------------------------- #
def freeze_below(ws: Worksheet, row: int, col: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row, column=col)


def hide_gridlines(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False


def label_value(
    ws: Worksheet, label: str, value: str, *, row: int, label_col: int = 1, value_col: int = 2,
    value_last_col: Optional[int] = None,
) -> int:
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = T.FONT_LABEL
    lc.alignment = T.ALIGN_LEFT_TOP
    if value_last_col and value_last_col > value_col:
        merge(ws, row, value_col, row, value_last_col)
    vc = ws.cell(row=row, column=value_col, value=value)
    vc.font = T.FONT_BODY
    vc.alignment = T.ALIGN_LEFT_TOP
    return row + 1
