"""Sheet 2 -- Executive Dashboard: KPI cards + native (editable) Excel charts.

Charts reference a small block of source data written to hidden far-right
columns so the visible canvas stays clean. Because they are real Excel chart
objects they recalculate/redraw automatically when the data changes.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.worksheet import Worksheet

from .. import components as C
from .. import theme as T
from ..dataset import ReportDataset

LAST_COL = 12
_SRC_COL = 1           # source data starts in col A of the hidden data sheet
_DATA_SHEET = "_ChartData"


def render(ws: Worksheet, ds: ReportDataset) -> None:
    C.hide_gridlines(ws)
    C.set_widths(ws, [16] * LAST_COL)
    k = ds.kpis

    row = C.title_band(
        ws, "AI Tool Usage Dashboard",
        f"Executive view · {ds.period.label} ({ds.period.days} days) · auto-generated, no manual interpretation needed",
        last_col=LAST_COL,
    )

    # ---- KPI cards: two rows of four ------------------------------------- #
    row = C.kpi_cards(ws, [
        C.Kpi("Total Employees", k.total_employees, T.FMT_INT),
        C.Kpi("Total Tools Subscribed", k.total_tools, T.FMT_INT),
        C.Kpi("Tools Integrated", k.tools_integrated, T.FMT_INT),
        C.Kpi("Employees Using AI", k.employees_using_ai, T.FMT_INT),
    ], row=row, span=3)
    row = C.kpi_cards(ws, [
        C.Kpi("Overall Adoption", k.adoption_pct, T.FMT_PCT),
        C.Kpi("Total Sessions", k.total_sessions, T.FMT_INT),
        C.Kpi("Total Generations", k.total_generations, T.FMT_INT),
        C.Kpi("Total Credits Burned", round(k.total_credits), T.FMT_INT),
    ], row=row, span=3)

    row = C.callout(
        ws,
        f"Overall AI adoption: {k.adoption_pct:.0%} of employees used at least one tracked AI tool this period.",
        row=row, last_col=LAST_COL,
    )

    anchor_row = C.section_header(ws, "Analytics", row=row, last_col=LAST_COL) + 1

    # Chart source data goes on a dedicated hidden worksheet. (Charts cannot plot
    # from fully-hidden *columns* on their own sheet — Excel drops the series —
    # but a hidden data *sheet* works reliably.)
    data_ws = ws.parent.create_sheet(_DATA_SHEET)
    starts = _write_sources(data_ws, ds)
    data_ws.sheet_state = "hidden"
    _build_charts(ws, data_ws, ds, anchor_row, starts)


# --------------------------------------------------------------------------- #
# Source data (hidden sheet) + charts
# --------------------------------------------------------------------------- #
def _write_sources(ws: Worksheet, ds: ReportDataset) -> dict[str, int]:
    r = 1
    starts: dict[str, int] = {}

    def block(name: str, header: list[str], rows: list[list]) -> None:
        nonlocal r
        starts[name] = r
        for j, h in enumerate(header):
            ws.cell(row=r, column=_SRC_COL + j, value=h)
        r += 1
        for data_row in rows:
            for j, v in enumerate(data_row):
                ws.cell(row=r, column=_SRC_COL + j, value=v)
            r += 1
        r += 1  # spacer

    block("tool", ["Tool", "Employees"], [[t.tool, t.employees_using] for t in ds.tool_usage])
    block("dept", ["Department", "Adoption %"], [[d.department, round(d.pct, 3)] for d in ds.dept_adoption])
    block("daily", ["Day", "ChatGPT", "Kling"],
          [[d.day.strftime("%d-%b"), d.chatgpt, d.kling] for d in ds.daily])
    block("top", ["Employee", "Composite"], [[e.name, e.composite_score] for e in ds.top_employees])
    block("dist", ["Category", "Employees"], _distribution(ds))
    return starts


def _distribution(ds: ReportDataset) -> list[list]:
    order = ["Inactive", "Light", "Moderate", "Heavy"]
    counts = {label: 0 for label in order}
    for emp in ds.employees:
        counts[emp.usage_category] = counts.get(emp.usage_category, 0) + 1
    return [[label, counts[label]] for label in order]


def _build_charts(chart_ws: Worksheet, src_ws: Worksheet, ds: ReportDataset,
                  anchor_row: int, src: dict[str, int]) -> None:
    # 1. Tool Adoption (column)
    n_tool = len(ds.tool_usage)
    bar = _bar("Tool Adoption — Employees Using Each Tool", "Employees")
    _add_bar_series(bar, src_ws, src["tool"], n_tool, color=T.NAVY)
    chart_ws.add_chart(bar, f"A{anchor_row}")

    # 2. Department Adoption (column, %)
    n_dept = len(ds.dept_adoption)
    dbar = _bar("Department Adoption Rate", "Adoption %")
    _add_bar_series(dbar, src_ws, src["dept"], n_dept, color=T.GREEN, pct=True)
    chart_ws.add_chart(dbar, f"G{anchor_row}")

    # 3. Daily Usage Trend (line, ChatGPT vs Kling)
    n_day = len(ds.daily)
    line = LineChart()
    line.title = "Daily Usage Trend"
    line.style = 2
    line.height, line.width = 7.5, 15
    data = Reference(src_ws, min_col=_SRC_COL + 1, max_col=_SRC_COL + 2,
                     min_row=src["daily"], max_row=src["daily"] + n_day)
    cats = Reference(src_ws, min_col=_SRC_COL, min_row=src["daily"] + 1, max_row=src["daily"] + n_day)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.title = "Events"
    line.x_axis.delete = False
    line.y_axis.delete = False
    chart_ws.add_chart(line, f"A{anchor_row + 16}")

    # 4. Top Employees (horizontal bar)
    n_top = len(ds.top_employees)
    tbar = _bar("Top Employees by AI Activity", "Composite score", horizontal=True)
    _add_bar_series(tbar, src_ws, src["top"], n_top, color="2E75B6")
    chart_ws.add_chart(tbar, f"G{anchor_row + 16}")

    # 5. AI Usage Distribution (pie)
    pie = PieChart()
    pie.title = "AI Usage Distribution"
    pie.height, pie.width = 7.5, 15
    pdata = Reference(src_ws, min_col=_SRC_COL + 1, min_row=src["dist"], max_row=src["dist"] + 4)
    pcats = Reference(src_ws, min_col=_SRC_COL, min_row=src["dist"] + 1, max_row=src["dist"] + 4)
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(pcats)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showCatName = True
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showLegendKey = False
    chart_ws.add_chart(pie, f"A{anchor_row + 32}")


def _bar(title: str, y_title: str, *, horizontal: bool = False) -> BarChart:
    chart = BarChart()
    chart.type = "bar" if horizontal else "col"
    chart.title = title
    chart.style = 10
    chart.height, chart.width = 7.5, 13
    chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None
    return chart


def _add_bar_series(chart: BarChart, ws: Worksheet, start: int, n: int, *, color: str, pct: bool = False) -> None:
    data = Reference(ws, min_col=_SRC_COL + 1, min_row=start, max_row=start + n)
    cats = Reference(ws, min_col=_SRC_COL, min_row=start + 1, max_row=start + n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = color
    if pct:
        chart.y_axis.numFmt = "0%"
        chart.y_axis.majorGridlines = None
