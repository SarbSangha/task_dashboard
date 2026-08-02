"""
Sheet -- Employee Summary with an inline, five-level expand/collapse drill-down.

Everything lives on this one sheet; nothing navigates away. Excel outline groups
give the +/- controls in the left margin:

    Level 1  Employee row .......... EMP-2026-0011  Devraj Singh  ...
    Level 2    └ Date row .......... 01-Jul-2026 · 52 events
    Level 3        └ Client row .... ⊞ Acme Corp · 20 gen · 640 credits
    Level 4            └ Tool row .. ⊞ Kling · 50 events · 1,495 credits
    Level 5                └ Log rows .. that day's Kling events for that employee/client

Client Mapping (the admin-curated Client list, GenerationClient) exists for
Freepik and Kling, the two providers with a pre-generation Task/Client picker
- see Event.client_name. ChatGPT's events, and any Freepik/Kling row nobody
linked to a client, share one "No Client" bucket at level 3, so the tree
stays navigable regardless of whether Client Mapping applies.

All detail levels are collapsed on open, so the sheet still reads as the plain
employee table until someone drills in.

Trade-off: because detail rows are interleaved between employee rows, this sheet
cannot also be an Excel Table (a ListObject must be a uniform rectangle), so the
header filter dropdowns are not available here. The ChatGPT / Kling Usage Log
sheets keep full filtering for that purpose.
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.properties import Outline
from openpyxl.worksheet.worksheet import Worksheet

from .. import components as C
from .. import theme as T
from ..dataset import Employee, Event, ReportDataset

LAST_COL = 18
_ADOPTION_COL = 16  # column P

# Guard rail: a single employee with a runaway period cannot explode the sheet.
# The full record always remains on the raw ChatGPT / Kling log sheets.
MAX_DETAIL_ROWS_PER_EMPLOYEE = 1500

_TOOL_TINT = {"ChatGPT": T.TINT_CHATGPT, "Kling": T.TINT_KLING, "Freepik": T.TINT_FREEPIK}

# Employee columns. Widths are a deliberate compromise: columns B/C also carry
# the prompt/response text of the nested log rows, so they are wider than a
# name/department alone would need.
_EMP_COLS = [
    C.Col("Employee ID", 15, "center", key="employee_id"),
    C.Col("Employee Name", 30, "left", key="name"),
    C.Col("Department", 30, "left", key="department"),
    C.Col("ChatGPT Sessions", 16, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.chatgpt_sessions),
    C.Col("ChatGPT Last Used", 16, "center", fmt=T.FMT_DATE, get=lambda e: e.chatgpt_last),
    C.Col("Kling Videos Made", 14, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.kling_videos),
    C.Col("Kling Credits Used", 15, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.kling_credits),
    C.Col("Kling Last Used", 24, "center", fmt=T.FMT_DATE, get=lambda e: e.kling_last),
    C.Col("Freepik Images", 14, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.freepik_images),
    C.Col("Freepik Videos", 14, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.freepik_videos),
    C.Col("Freepik Credits Charged", 17, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.freepik_credits_charged),
    C.Col("Freepik Credits Estimated", 18, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.freepik_credits_estimated),
    C.Col("Freepik Last Used", 24, "center", fmt=T.FMT_DATE, get=lambda e: e.freepik_last),
    C.Col("Total AI Usage", 13, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.total_usage),
    C.Col("Tools Used", 11, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.tools_used),
    C.Col("Adoption Status", 20, "center", get=lambda e: e.adoption_status),
    C.Col("Composite Score", 14, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.composite_score),
    C.Col("Usage Category", 14, "center", get=lambda e: e.usage_category),
]

# Nested log columns deliberately line up with the employee columns above them:
#   F = Videos  -> "Kling Videos Made"      G = Credits -> "Kling Credits Used"
#   H = Ref ID  -> "Kling Last Used"
# E used to be a blank spacer; it now shows each row's actual output type
# (Image/Video) - Freepik is mixed-media (unlike Kling, which is always a
# video, and ChatGPT, which is text), so without this a Freepik image row and
# a Freepik video row were indistinguishable at a glance.
_LOG_COLS = [
    C.Col("Date", 15, "center", fmt=T.FMT_DATE, get=lambda e: e.when),      # A
    C.Col("Prompt / Input", 30, "left", key="prompt"),                      # B
    C.Col("Response / Output", 30, "left", key="response"),                 # C
    C.Col("Model", 16, "left", key="model"),                                # D
    C.Col("Type", 10, "center", key="kind"),                                # E
    C.Col("Videos", 14, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.videos),    # F
    C.Col("Credits", 15, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.credits),  # G
]

# Column indexes used to align the date/tool summary bands with the header.
_COL_CHATGPT_COUNT = 4   # D — "ChatGPT Sessions"
_COL_KLING_COUNT = 6     # F — "Kling Videos Made"
_COL_KLING_CREDITS = 7   # G — "Kling Credits Used"
_COL_FREEPIK_IMAGES = 9    # I — "Freepik Images"
_COL_FREEPIK_VIDEOS = 10   # J — "Freepik Videos"
_COL_FREEPIK_CREDITS = 11  # K — "Freepik Credits Charged"


def render(ws: Worksheet, ds: ReportDataset) -> None:
    C.hide_gridlines(ws)
    # Summary rows sit ABOVE their detail, so each +/- lines up with the row clicked.
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

    row = C.title_band(
        ws, f"Employee Summary — All {len(ds.employees)} Employees (auto-calculated)",
        "Click the + in the left margin to expand an employee → date → tool → that day's log. "
        "Every employee appears here whether or not they used any AI tool.",
        last_col=LAST_COL,
    )

    header_row = row
    row = C.table_header(ws, _EMP_COLS, row=row)
    index = _index_events(ds.merged_events)

    for i, emp in enumerate(ds.employees):
        base = T.FILL_ALT if i % 2 else T.FILL_WHITE
        C.table_row(ws, _EMP_COLS, emp, row_idx=row, fill=base)
        ws.row_dimensions[row].outlineLevel = 0
        emp_row = row
        row += 1
        next_row = _render_detail(ws, emp, index.get(emp.employee_id, {}), row)
        if next_row > row:
            _collapse(ws, emp_row)  # employee row summarises its date groups
        row = next_row

    C.freeze_below(ws, header_row + 1, col=2)
    _apply_adoption_formatting(ws, header_row + 1, row - 1)

    # Filter dropdowns on the header. A plain AutoFilter (unlike an Excel Table)
    # tolerates the interleaved drill-down rows, so filtering and the +/- outline
    # coexist on this sheet.
    ws.auto_filter.ref = f"A{header_row}:{C.col_letter(LAST_COL)}{max(row - 1, header_row)}"


# --------------------------------------------------------------------------- #
# Detail levels
# --------------------------------------------------------------------------- #
def _index_events(events: list[Event]) -> dict:
    """employee_id -> date -> client -> tool -> [events], built in one pass.

    Client Mapping exists for Freepik and Kling (Event.client_name is blank
    for ChatGPT) - everything without a linked client, from any provider,
    shares the "No Client" bucket so the tree stays navigable."""
    index: dict = {}
    for ev in events:
        client_key = ev.client_name or "No Client"
        (
            index.setdefault(ev.employee_id, {})
            .setdefault(ev.when, {})
            .setdefault(client_key, {})
            .setdefault(ev.tool, [])
            .append(ev)
        )
    return index


def _flatten_tools(by_client: dict) -> dict:
    """Merge every client's tool->events map for one date into a single
    tool->events map - the date band's aggregate counts must reflect every
    client that day, not just whichever one is being rendered."""
    merged: dict = {}
    for by_tool in by_client.values():
        for tool, events in by_tool.items():
            merged.setdefault(tool, []).extend(events)
    return merged


def _render_detail(ws: Worksheet, emp: Employee, by_date: dict, row: int) -> int:
    """Render date -> client -> tool -> log rows beneath one employee. Returns next row."""
    if not by_date:
        return row

    budget = MAX_DETAIL_ROWS_PER_EMPLOYEE
    for day in sorted(d for d in by_date if d is not None):
        by_client = by_date[day]

        # Level 2 — the date. Counts sit under the same headers as the employee
        # row above: ChatGPT under D, Kling under F/G, Freepik under I/J -
        # aggregated across every client that day (see _flatten_tools).
        date_row = row
        _band_row(ws, row, level=1, fill=T.solid(T.KPI_BAND), indent=1,
                  cells=_count_cells(day, _flatten_tools(by_client)))
        row += 1

        for client_name in sorted(by_client):
            by_tool = by_client[client_name]

            # Level 3 — the client for that date
            client_row = row
            _band_row(ws, row, level=2, fill=T.solid(T.TINT_CLIENT), indent=2,
                      cells=_count_cells(f"⊞  {client_name}", by_tool, label_fmt=None))
            row += 1

            for tool in sorted(by_tool):
                events = by_tool[tool]
                tint = T.solid(_TOOL_TINT.get(tool, T.ALT_ROW))

                # Level 4 — the tool for that client/date
                tool_row = row
                _band_row(ws, row, level=3, fill=tint, indent=3,
                          cells=_count_cells(f"⊞  {tool}", {tool: events}, label_fmt=None))
                row += 1

                # Level 5 — that day's log for that client + tool
                shown = events[:budget]
                C.table_header(ws, _LOG_COLS, row=row, apply_widths=False)
                _set_level(ws, row, 4)
                row += 1
                for ev in shown:
                    C.table_row(ws, _LOG_COLS, ev, row_idx=row, fill=tint)
                    _set_level(ws, row, 4)
                    row += 1
                budget -= len(shown)

                if len(events) > len(shown):
                    note = ws.cell(row=row, column=1,
                                   value=f"… {len(events) - len(shown):,} more {tool} event(s) on this date are not "
                                         f"listed (per-employee detail row cap reached). The counts above remain complete.")
                    note.font = T.FONT_NOTE
                    C.merge(ws, row, 1, row, LAST_COL)
                    _set_level(ws, row, 4)
                    row += 1

                _collapse(ws, tool_row)  # tool row summarises its log rows
                if budget <= 0:
                    _collapse(ws, client_row)
                    _collapse(ws, date_row)
                    return row

            _collapse(ws, client_row)  # client row summarises its tool groups

        _collapse(ws, date_row)  # date row summarises its client groups
    return row


def _count_cells(label, by_tool: dict, label_fmt=T.FMT_DATE) -> list:
    """Build band cells: label in A, per-tool counts/credits under their own
    headers - computed straight from ``by_tool`` rather than a single passed-in
    total, so a mixed (date-level) band with both Kling and Freepik activity
    that day attributes each tool's credits to its own column instead of
    conflating them into one number."""
    cells = [(1, label, label_fmt, "left")]
    chatgpt = len(by_tool.get("ChatGPT", ()))
    kling = len(by_tool.get("Kling", ()))
    freepik = len(by_tool.get("Freepik", ()))
    if chatgpt:
        cells.append((_COL_CHATGPT_COUNT, chatgpt, T.FMT_INT, "right"))
    if kling:
        cells.append((_COL_KLING_COUNT, kling, T.FMT_INT, "right"))
        kling_credits = sum((ev.credits or 0) for ev in by_tool["Kling"])
        if kling_credits:
            cells.append((_COL_KLING_CREDITS, round(kling_credits), T.FMT_INT, "right"))
    if freepik:
        freepik_events = by_tool["Freepik"]
        freepik_videos = sum(1 for ev in freepik_events if ev.kind == "Video")
        freepik_images = freepik - freepik_videos
        if freepik_images:
            cells.append((_COL_FREEPIK_IMAGES, freepik_images, T.FMT_INT, "right"))
        if freepik_videos:
            cells.append((_COL_FREEPIK_VIDEOS, freepik_videos, T.FMT_INT, "right"))
        freepik_credits = sum((ev.credits or 0) for ev in freepik_events)
        if freepik_credits:
            cells.append((_COL_FREEPIK_CREDITS, round(freepik_credits), T.FMT_INT, "right"))
    return cells


def _band_row(ws: Worksheet, row: int, *, level: int, fill, indent: int, cells) -> None:
    """A summary band (date or tool) spanning the employee columns."""
    for col in range(1, LAST_COL + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = T.BORDER_CELL
    for col, value, fmt, align in cells:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = T.FONT_BODY_BOLD
        cell.alignment = (
            Alignment(horizontal="left", vertical="center", indent=indent)
            if align == "left"
            else Alignment(horizontal="right", vertical="center")
        )
        if fmt:
            cell.number_format = fmt
    ws.row_dimensions[row].height = 18
    _set_level(ws, row, level)


def _set_level(ws: Worksheet, row: int, level: int) -> None:
    """Put a row in an outline group and hide it (collapsed by default)."""
    dim = ws.row_dimensions[row]
    dim.outlineLevel = level
    dim.hidden = True


def _collapse(ws: Worksheet, row: int) -> None:
    """Mark a summary row as collapsed.

    Hiding the detail rows alone is not enough: Excel drives the +/- control from
    the ``collapsed`` flag on the summary row. Without it the outline state is
    inconsistent (children hidden but the control still says "expanded") and
    Excel re-expands the group on open.
    """
    ws.row_dimensions[row].collapsed = True


# --------------------------------------------------------------------------- #
def _apply_adoption_formatting(ws: Worksheet, first_row: int, last_row: int) -> None:
    """Colour the Adoption Status column by value.

    Applied across the whole used range: nested detail rows never contain these
    exact strings, so only real employee rows are affected.
    """
    if last_row < first_row:
        return
    col = C.col_letter(_ADOPTION_COL)
    rng = f"{col}{first_row}:{col}{last_row}"
    for value, fill, colour in (
        ("Not Used", T.FILL_NOT_USED, T.TEXT_NOT_USED),
        ("Using Multiple Tools", T.FILL_MULTI, T.TEXT_MULTI),
        ("Using 1 Tool", T.FILL_ONE, T.TEXT_ONE),
    ):
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{value}"'],
                       fill=PatternFill("solid", fgColor=fill),
                       font=Font(name=T.FONT_NAME, size=10, color=colour)),
        )
