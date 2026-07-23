"""Sheet 1 -- Read Me: purpose, legend, metadata and data-quality panel."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from .. import components as C
from .. import theme as T
from ..dataset import ReportDataset

LAST_COL = 6

_PURPOSE = (
    "This workbook is the fortnightly executive record of AI-tool adoption and "
    "usage across the organization. It is generated directly from the AI "
    "Dashboard's live capture database — every figure is derived, never typed."
)

_DESCRIPTION = (
    "It consolidates ChatGPT and Kling activity (with fifteen further tools on "
    "the roadmap) into one auditable source for leadership, AI governance, HR "
    "and operations: who is using AI, how much, on which tools, and where the "
    "adoption gaps are."
)

_LEGEND = [
    ("Adoption Status", "Not Used = 0 tools · Using 1 Tool · Using Multiple Tools = 2+ tools in the period."),
    ("Composite Score", "ChatGPT sessions + Kling videos for the employee this period."),
    ("Employee drill-down", "On Employee Summary, click a highlighted employee name to open their date-wise ChatGPT + Kling activity log."),
    ("ChatGPT Session", "One captured prompt (counted from the conversation prompt counter)."),
    ("Kling Generation", "One Kling usage event — the actual generation performed on the tool."),
    ("Credits", "Kling platform credits consumed (the real spend signal)."),
    ("— (em dash)", "No activity / not captured for this field."),
]

_NOTES = [
    "Fully auto-generated: no cell is hand-edited. Re-running replaces every sheet.",
    "Covers the selected reporting period (see Report Metadata); the default is a rolling 15-day cycle.",
    "On Employee Summary, use the + controls in the left margin to expand employee → date → tool → that day's log.",
    "Employee Summary has filter dropdowns on its header row; totals always reflect the full period.",
    "Adding a new AI tool (Claude, Gemini, …) lights up automatically; no layout change needed.",
]


def render(ws: Worksheet, ds: ReportDataset) -> None:
    C.hide_gridlines(ws)
    C.set_widths(ws, [22, 20, 20, 20, 20, 20])

    row = C.title_band(
        ws, "AI Tool Usage Report — Read Me",
        "How to read this workbook · what it covers · how it is generated",
        last_col=LAST_COL,
    )

    row = C.section_header(ws, "Purpose", row=row, last_col=LAST_COL)
    row = _paragraph(ws, _PURPOSE, row)
    row = _paragraph(ws, _DESCRIPTION, row)
    row += 1

    row = C.section_header(ws, "Legend", row=row, last_col=LAST_COL)
    for term, meaning in _LEGEND:
        row = _term_line(ws, term, meaning, row)
    row += 1

    row = C.section_header(ws, "Auto-Generation Notes", row=row, last_col=LAST_COL)
    for note in _NOTES:
        row = _bullet(ws, note, row)
    row += 1

    row = C.section_header(ws, "Report Metadata", row=row, last_col=LAST_COL)
    meta = [
        ("Report period", ds.period.label),
        ("Last generated", ds.generated_at.strftime("%d-%b-%Y %H:%M")),
        ("Version", ds.version),
        ("Data source", "AI Dashboard capture database (ChatGPT + Kling pipelines)"),
        ("Employees covered", f"{ds.kpis.total_employees:,}"),
        ("Tools subscribed / integrated", f"{ds.kpis.total_tools} / {ds.kpis.tools_integrated}"),
        ("Events this period", f"{ds.kpis.total_sessions + ds.kpis.total_generations:,} "
                              f"({ds.kpis.total_sessions:,} ChatGPT · {ds.kpis.total_generations:,} Kling)"),
    ]
    for label, value in meta:
        row = C.label_value(ws, label, value, row=row, value_col=2, value_last_col=LAST_COL)
    row += 1

    row = C.section_header(ws, "Data-Quality Checks", row=row, last_col=LAST_COL)
    cols = [
        C.Col("Severity", 14, "center", key="severity"),
        C.Col("Check", 26, "left", key="check"),
        C.Col("Detail", 78, "left", get=lambda w: w.detail),
    ]

    def sev_fill(_i, w):
        if w.severity == "Error":
            return T.solid(T.FILL_NOT_USED)
        if w.severity == "Warning":
            return T.solid(T.FILL_PENDING)
        return None

    C.data_table(ws, cols, ds.warnings, start_row=row, row_fill=sev_fill)


# --------------------------------------------------------------------------- #
def _paragraph(ws: Worksheet, text: str, row: int) -> int:
    C.merge(ws, row, 1, row, LAST_COL)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = T.FONT_BODY
    cell.alignment = T.ALIGN_LEFT_TOP
    ws.row_dimensions[row].height = 46
    return row + 1


def _term_line(ws: Worksheet, term: str, meaning: str, row: int) -> int:
    tc = ws.cell(row=row, column=1, value=term)
    tc.font = T.FONT_BODY_BOLD
    tc.alignment = T.ALIGN_LEFT_TOP
    C.merge(ws, row, 2, row, LAST_COL)
    mc = ws.cell(row=row, column=2, value=meaning)
    mc.font = T.FONT_BODY
    mc.alignment = T.ALIGN_LEFT_TOP
    return row + 1


def _bullet(ws: Worksheet, text: str, row: int) -> int:
    C.merge(ws, row, 1, row, LAST_COL)
    cell = ws.cell(row=row, column=1, value=f"•  {text}")
    cell.font = T.FONT_BODY
    cell.alignment = T.ALIGN_LEFT_TOP
    return row + 1
