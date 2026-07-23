"""
Per-employee drill-down sheet: tools listed row by row, each expanding to its log.

Two levels, both on one page:
  1. "Tools Used" lists every tool the employee touched (events, credits, last used).
  2. Each tool row owns a collapsed Excel *outline group* holding that tool's own
     log for that employee. Clicking the ⊞ in the left margin (or the outline
     "2" button) drops the log open in place, like a dropdown — no macros needed.

Groups are collapsed on open so the sheet reads as a short tool list first.
"""

from __future__ import annotations

from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.properties import Outline
from openpyxl.worksheet.worksheet import Worksheet

from .. import components as C
from .. import theme as T
from ..dataset import Employee, Event, ReportDataset

LAST_COL = 8
SUMMARY_SHEET = "Employee Summary"

_BAD_CHARS = set(r"[]:*?/\'")

# Shared column layout for every tool's log, so all groups line up on the sheet.
_LOG_COLS = [
    C.Col("Date", 13, "center", fmt=T.FMT_DATE, get=lambda e: e.when),
    C.Col("Prompt / Input", 42, "left", key="prompt"),
    C.Col("Response / Output", 30, "left", key="response"),
    C.Col("Model", 15, "left", key="model"),
    C.Col("Credits", 10, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.credits),
    C.Col("Videos", 8, "right", fmt=T.FMT_INT_DASH, get=lambda e: e.videos),
    C.Col("Status", 12, "center", key="status"),
    C.Col("Reference ID", 24, "left", key="ref_id"),
]

_TOOL_TINT = {"ChatGPT": T.TINT_CHATGPT, "Kling": T.TINT_KLING}


def sheet_name_for(emp: Employee) -> str:
    """Stable, Excel-legal sheet name for an employee's drill-down."""
    raw = (emp.employee_id or f"U{emp.user_id}").strip()
    cleaned = "".join(ch for ch in raw if ch not in _BAD_CHARS)
    return cleaned[:31] or f"U{emp.user_id}"


def render(ws: Worksheet, ds: ReportDataset, emp: Employee, events: list[Event]) -> None:
    C.hide_gridlines(ws)
    # Summary row sits ABOVE its detail group, so the +/- control lines up with
    # the tool row the user clicks.
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

    row = C.title_band(
        ws, f"{emp.name} — AI Activity Log",
        f"{emp.employee_id} · {emp.department} · {ds.period.label}",
        last_col=LAST_COL,
    )

    back = ws.cell(row=row, column=1, value="← Back to Employee Summary")
    back.font = T.FONT_LINK
    back.hyperlink = Hyperlink(ref=back.coordinate, location=f"'{SUMMARY_SHEET}'!A1")
    row += 2

    row = C.kpi_cards(ws, [
        C.Kpi("ChatGPT Sessions", emp.chatgpt_sessions, T.FMT_INT),
        C.Kpi("Kling Generations", emp.kling_videos, T.FMT_INT),
        C.Kpi("Kling Credits Used", round(emp.kling_credits), T.FMT_INT),
        C.Kpi("Total AI Usage", emp.total_usage, T.FMT_INT),
    ], row=row, span=2)

    row = C.section_header(
        ws, "Tools Used — click the ⊞ in the left margin to open that tool's log",
        row=row, last_col=LAST_COL,
    )
    row = _tools_header(ws, row)

    # One block per tool the employee actually used.
    for tool, count, credits, last_used in (
        ("ChatGPT", emp.chatgpt_sessions, None, emp.chatgpt_last),
        ("Kling", emp.kling_videos, emp.kling_credits, emp.kling_last),
    ):
        if not count:
            continue
        tool_events = [ev for ev in events if ev.tool == tool]
        row = _tool_block(ws, row, tool, count, credits, last_used, tool_events)

    if not emp.total_usage:
        note = ws.cell(row=row, column=1, value="This employee recorded no AI tool usage in the selected period.")
        note.font = T.FONT_NOTE


# --------------------------------------------------------------------------- #
def _tools_header(ws: Worksheet, row: int) -> int:
    for j, (text, width) in enumerate(
        [("Tool", 22), ("Events", 12), ("Credits Used", 14), ("Last Used", 14)]
    ):
        cell = ws.cell(row=row, column=1 + j, value=text)
        cell.font = T.FONT_HEADER
        cell.fill = T.FILL_HEADER
        cell.alignment = T.ALIGN_CENTER
        cell.border = T.BORDER_HEADER
        # Widths are also set by the log tables; keep the first four readable.
        ws.column_dimensions[C.col_letter(1 + j)].width = max(
            width, ws.column_dimensions[C.col_letter(1 + j)].width or 0
        )
    ws.row_dimensions[row].height = 22
    return row + 1


def _tool_block(ws: Worksheet, row: int, tool: str, count: int,
                credits, last_used, events: list[Event]) -> int:
    """A visible tool summary row + its collapsed log group underneath."""
    fill = T.solid(_TOOL_TINT.get(tool, T.ALT_ROW))
    values = [f"⊞  {tool}", count, credits if credits is not None else "—", last_used]
    for j, value in enumerate(values):
        cell = ws.cell(row=row, column=1 + j, value=value)
        cell.font = T.FONT_BODY_BOLD
        cell.fill = fill
        cell.border = T.BORDER_CELL
        cell.alignment = T.ALIGN_LEFT if j == 0 else T.ALIGN_RIGHT if j in (1, 2) else T.ALIGN_CENTER
        if j == 1:
            cell.number_format = T.FMT_INT
        elif j == 2 and credits is not None:
            cell.number_format = T.FMT_INT
        elif j == 3:
            cell.number_format = T.FMT_DATE
    ws.row_dimensions[row].height = 20
    detail_start = row + 1

    if events:
        end = C.data_table(
            ws, _LOG_COLS, events, start_row=detail_start,
            table_name=None, autofilter=False,
            row_fill=lambda _i, _ev, f=fill: f,
        )
        last_detail = end - 2
    else:
        # The tool was used but no row-level detail was captured — say so plainly
        # rather than showing an empty expander.
        cell = ws.cell(
            row=detail_start, column=1,
            value=f"{count:,} {tool} event(s) counted; row-level detail is not available for this period.",
        )
        cell.font = T.FONT_NOTE
        C.merge(ws, detail_start, 1, detail_start, LAST_COL)
        last_detail = detail_start

    # Collapse the detail rows into an outline group under the tool row.
    for r in range(detail_start, last_detail + 1):
        ws.row_dimensions[r].outlineLevel = 1
        ws.row_dimensions[r].hidden = True

    return last_detail + 2  # spacer before the next tool
