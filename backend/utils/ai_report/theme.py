"""
The design system for the AI Usage Workbook.

Every colour, font, fill, border and number format the workbook uses is defined
here exactly once. Sheet modules never hard-code a hex value or a font size --
they import from this module. That keeps the seven sheets visually identical to
the approved screenshots and makes a global restyle a one-file change.

Palette was lifted directly from the reference screenshots:
  * Navy title bands            -> #1F3864
  * Navy table column headers    -> #1F4E79
  * Green section headers        -> #548235 / #375623
  * Light-blue KPI card band     -> #D9E1F2
  * Zebra / alternating rows     -> #F2F2F2
  * Status tints (adoption)      -> salmon #FCE4D6 / green #E2EFDA
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# --------------------------------------------------------------------------- #
# Colours (ARGB hex, no leading '#')
# --------------------------------------------------------------------------- #
NAVY = "1F3864"            # title bands
NAVY_TABLE = "1F4E79"      # table column headers
GREEN = "548235"           # section headers
GREEN_DARK = "375623"      # section header text / accents
KPI_BAND = "D9E1F2"        # light-blue KPI card background
WHITE = "FFFFFF"
BLACK = "000000"

TEXT_DARK = "1F3864"       # primary numbers / emphasis
TEXT_BODY = "262626"       # regular body text
TEXT_MUTED = "595959"      # labels, subtitles
TEXT_SUBTLE = "808080"     # legend / footnotes

ALT_ROW = "F2F2F2"         # zebra striping
GRID = "D9D9D9"            # thin cell borders
BAND_RULE = "BFBFBF"       # heavier separators

# Provider row tints used in the merged Overview log.
TINT_CHATGPT = "F2F6FC"    # very light blue
TINT_KLING = "EBF1DE"      # very light olive/green

# Status / conditional-format tints (adoption + integration state).
FILL_NOT_USED = "FCE4D6"
TEXT_NOT_USED = "843C0C"
FILL_MULTI = "E2EFDA"
TEXT_MULTI = "375623"
FILL_ONE = "FFF2CC"
TEXT_ONE = "7F6000"
FILL_INTEGRATED = "E2EFDA"
TEXT_INTEGRATED = "375623"
FILL_PENDING = "FFF2CC"
TEXT_PENDING = "BF8F00"

FONT_NAME = "Calibri"

# --------------------------------------------------------------------------- #
# Fills
# --------------------------------------------------------------------------- #
def solid(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color, bgColor=color)


FILL_TITLE = solid(NAVY)
FILL_SECTION = solid(GREEN)
FILL_KPI = solid(KPI_BAND)
FILL_HEADER = solid(NAVY_TABLE)
FILL_ALT = solid(ALT_ROW)
FILL_WHITE = solid(WHITE)

# --------------------------------------------------------------------------- #
# Fonts
# --------------------------------------------------------------------------- #
FONT_TITLE = Font(name=FONT_NAME, size=16, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name=FONT_NAME, size=9, italic=True, color=TEXT_SUBTLE)
FONT_SECTION = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
FONT_KPI_LABEL = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_MUTED)
FONT_KPI_VALUE = Font(name=FONT_NAME, size=24, bold=True, color=NAVY)
FONT_KPI_SUB = Font(name=FONT_NAME, size=9, color=TEXT_SUBTLE)
FONT_HEADER = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
FONT_BODY = Font(name=FONT_NAME, size=10, color=TEXT_BODY)
FONT_BODY_BOLD = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_BODY)
FONT_NUM = Font(name=FONT_NAME, size=10, color=TEXT_BODY)
FONT_CALLOUT = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
FONT_LABEL = Font(name=FONT_NAME, size=10, bold=True, color=TEXT_MUTED)
FONT_NOTE = Font(name=FONT_NAME, size=9, color=TEXT_MUTED)

# --------------------------------------------------------------------------- #
# Alignment
# --------------------------------------------------------------------------- #
ALIGN_TITLE = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_KPI_LABEL = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_KPI_VALUE = Alignment(horizontal="center", vertical="center")

# --------------------------------------------------------------------------- #
# Borders
# --------------------------------------------------------------------------- #
_thin = Side(style="thin", color=GRID)
_medium = Side(style="medium", color=WHITE)
BORDER_CELL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_HEADER = Border(
    left=Side(style="thin", color=WHITE),
    right=Side(style="thin", color=WHITE),
    top=_thin,
    bottom=_thin,
)
BORDER_NONE = Border()

# --------------------------------------------------------------------------- #
# Number formats
# --------------------------------------------------------------------------- #
FMT_DATE = "DD-Mmm-YYYY"
FMT_INT = "#,##0"
FMT_INT_DASH = "#,##0;-#,##0;\"—\""   # show em-dash for zero
FMT_PCT = "0%"
FMT_PCT1 = "0.0%"
FMT_DECIMAL1 = "#,##0.0"
FMT_MONEY = "₹#,##0"             # INR
FMT_MONEY2 = "₹#,##0.00"

# Excel table style used for every ListObject (banded rows + light border).
TABLE_STYLE = "TableStyleLight1"
