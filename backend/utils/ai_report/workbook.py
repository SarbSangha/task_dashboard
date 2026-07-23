"""
Workbook orchestrator: build the dataset, then render the seven sheets in order.

This is the module the API/CLI imports. It owns sheet order, tab colours and the
final byte serialization -- nothing else. Each sheet's *content* lives in its own
module under ``sheets/``.
"""

from __future__ import annotations

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from sqlalchemy.orm import Session

from . import components as _components
from . import theme as T
from .dataset import ReportDataset, build_dataset
from .sheets import (  # noqa: F401 — disabled sheets stay imported for easy re-enabling
    chatgpt_log,     # kept for re-enabling; not currently rendered
    dashboard,
    employee_summary,
    kling_log,       # kept for re-enabling; not currently rendered
    overview,        # kept for re-enabling; not currently rendered
    read_me,
    tool_master,     # kept for re-enabling; not currently rendered
    user_log,        # kept for re-enabling; drill-down is now inline
)

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# (sheet title, renderer module, tab colour). Order == tab order.
#
# NOTE: "Overview (Summary + Raw Data)" and "Tool Master" are intentionally NOT
# generated right now (business decision). Their renderer modules are kept fully
# intact and still imported, so re-enabling them is a one-line uncomment below --
# no code was deleted.
_SHEETS = [
    ("Read Me", read_me, T.NAVY),
    ("Dashboard", dashboard, T.NAVY),
    # ("Overview (Summary + Raw Data)", overview, T.GREEN),   # disabled — see note above
    # ("Tool Master", tool_master, T.NAVY),                   # disabled — see note above
    ("Employee Summary", employee_summary, T.GREEN),
    # ("ChatGPT Usage Log", chatgpt_log, T.NAVY_TABLE),       # disabled — see note above
    # ("Kling Usage Log", kling_log, T.NAVY_TABLE),           # disabled — see note above
]


def render_workbook(ds: ReportDataset) -> Workbook:
    """Render a fully-populated workbook from an already-built dataset."""
    # Table names must be unique across the whole workbook; reset the guard so
    # repeated builds in one process (e.g. a scheduler) stay clean.
    _components._used_table_names.clear()

    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; we create our own

    for title, module, tab in _SHEETS:
        ws = wb.create_sheet(title=title[:31])  # Excel caps sheet names at 31 chars
        ws.sheet_properties.tabColor = tab
        module.render(ws, ds)

    # Per-employee drill-down now lives INLINE on the Employee Summary sheet
    # (user -> date -> tool -> log, via outline groups), so the separate
    # per-user sheets are no longer generated. Kept for easy re-enabling.
    # _render_user_sheets(wb, ds)

    wb.properties.title = "AI Tool Usage Report"
    wb.properties.creator = "AI Dashboard — Automated Reporting"
    wb.properties.subject = f"AI adoption & usage · {ds.period.label}"
    return wb


def _render_user_sheets(wb: Workbook, ds: ReportDataset) -> None:
    """One drill-down sheet per employee with activity, linked from the summary.

    Events are bucketed by employee in a single pass so this stays O(n) in the
    number of events rather than re-scanning the log for every employee.
    """
    by_employee: dict[str, list] = {}
    for ev in ds.merged_events:
        by_employee.setdefault(ev.employee_id, []).append(ev)

    for emp in ds.employees:
        if not emp.total_usage:
            continue  # nothing to drill into; name stays plain text on the summary
        ws = wb.create_sheet(title=user_log.sheet_name_for(emp))
        ws.sheet_properties.tabColor = T.NAVY_TABLE
        user_log.render(ws, ds, emp, by_employee.get(emp.employee_id, []))


def build_ai_workbook(
    db: Session,
    *,
    start: Optional[date] = None,
    end: Optional[date] = None,
    ref_date: Optional[date] = None,
) -> tuple[bytes, str, str]:
    """
    Public entry point. Returns ``(xlsx_bytes, mimetype, filename)``.

    A single call reads the DB, assembles the typed snapshot and serializes the
    workbook to bytes ready for a streaming HTTP response. See
    :func:`dataset.build_dataset` for how ``start`` / ``end`` / ``ref_date``
    resolve the reporting window.
    """
    ds = build_dataset(db, start=start, end=end, ref_date=ref_date)
    wb = render_workbook(ds)
    buf = io.BytesIO()
    wb.save(buf)
    filename = f"AI-Usage-Report_{ds.period.end:%Y-%m-%d}.xlsx"
    return buf.getvalue(), XLSX_MIMETYPE, filename
