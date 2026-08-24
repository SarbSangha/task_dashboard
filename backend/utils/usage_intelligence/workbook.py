"""
Usage Intelligence workbook.

Organisation/team downloads keep the 13-sheet pack. An individual download is
one visible tab a team lead can walk through: who they are, headline numbers,
plain-English summary, charts, and what to discuss. Prompt text is never written.
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter, FilterColumn, Filters

from utils.ai_report import components as C
from utils.ai_report import theme as T

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
LAST = 12
NOT_SET = "Not set"


def _blank(value, fallback=NOT_SET) -> str:
    if value is None:
        return fallback
    text = str(value).strip()
    if not text or text in {"—", "–", "-"}:
        return fallback
    return text


def _finish_axes(chart, *, x_title=None, y_title=None) -> None:
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title


def build_usage_workbook(snapshot: dict) -> tuple[bytes, str, str]:
    C._used_table_names.clear()
    period = snapshot["period"]
    individual = snapshot.get("individual")
    report_type = (snapshot.get("reportType") or "").strip().lower()

    if individual or report_type == "individual":
        person = individual or ((snapshot.get("users") or [None])[0])
        if person:
            return _build_person_workbook(snapshot, person)

    if report_type == "consolidated":
        return _build_consolidated_workbook(snapshot)

    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Executive Dashboard", _sheet_dashboard, T.NAVY),
        ("Executive Summary", _sheet_summary, T.GREEN),
        ("Team Overview", _sheet_teams, T.NAVY_TABLE),
        ("User Overview", _sheet_users, T.NAVY_TABLE),
        ("Individual Users", _sheet_individual, T.GREEN),
        ("Tool Usage", _sheet_tools, T.NAVY_TABLE),
        ("Consolidated", _sheet_consolidated, T.GREEN),
        ("Credit Analysis", _sheet_credits, T.NAVY),
        ("Generation Analysis", _sheet_generations, T.GREEN),
        ("Activity Timeline", _sheet_timeline, T.NAVY_TABLE),
        ("Trends", _sheet_trends, T.NAVY),
        ("Anomalies & Alerts", _sheet_anomalies, T.GREEN),
        ("Management Actions", _sheet_actions, T.NAVY),
        ("Raw Data", _sheet_raw, T.NAVY_TABLE),
    ]
    for title, renderer, color in sheets:
        ws = wb.create_sheet(title=title[:31])
        ws.sheet_properties.tabColor = color
        renderer(ws, snapshot)

    period = snapshot["period"]
    wb.properties.title = "Usage Intelligence Report"
    wb.properties.creator = "Task Dashboard — Usage Intelligence"
    wb.properties.subject = f"Management usage report · {period['label']}"

    buf = io.BytesIO()
    wb.save(buf)
    scope = snapshot.get("reportType") or "organisation"
    filename = f"Usage-Intelligence_{scope}_{period['end']}.xlsx"
    return buf.getvalue(), XLSX_MIMETYPE, filename


def _sheet_dashboard(ws, snapshot):
    C.hide_gridlines(ws)
    C.set_widths(ws, [16] * LAST)
    k = snapshot["kpis"]
    period = snapshot["period"]
    row = C.title_band(
        ws, "Usage Intelligence — Executive Dashboard",
        f"What is happening in the organisation · {period['label']} · vs previous {period['days']} days",
        last_col=LAST,
    )
    row = C.kpi_cards(ws, [
        C.Kpi("Active Users", k["activeUsers"]["value"], T.FMT_INT, sub=_delta(k["activeUsers"])),
        C.Kpi("Usage Hours", k["usageHours"]["value"], T.FMT_DECIMAL1, sub=_delta(k["usageHours"])),
        C.Kpi("Generations", k["generations"]["value"], T.FMT_INT, sub=_delta(k["generations"])),
        C.Kpi("Credits Consumed", k["credits"]["value"], T.FMT_INT, sub=_delta(k["credits"])),
    ], row=row, span=3)
    row = C.kpi_cards(ws, [
        C.Kpi("Success Rate", (k["successRate"]["value"] or 0) / 100.0, T.FMT_PCT1),
        C.Kpi("Most Used Tool", k["mostUsedTool"]),
        C.Kpi("Highest Credit Consumer", k["highestCreditConsumer"]),
        C.Kpi("Fastest Growing Tool", k["fastestGrowingTool"]),
    ], row=row, span=3)

    findings = snapshot.get("findings") or []
    if findings:
        row = C.callout(ws, "  ·  ".join(findings[:3]), row=row, last_col=LAST)

    # Hidden source for charts
    daily = (snapshot.get("trends") or {}).get("daily") or []
    src = ws
    src_col = 40
    src.cell(row=1, column=src_col, value="Date")
    src.cell(row=1, column=src_col + 1, value="Generations")
    src.cell(row=1, column=src_col + 2, value="Credits")
    for i, d in enumerate(daily, start=2):
        src.cell(row=i, column=src_col, value=d["date"])
        src.cell(row=i, column=src_col + 1, value=d["generations"])
        src.cell(row=i, column=src_col + 2, value=d["credits"])
    ws.column_dimensions[get_column_letter(src_col)].hidden = True
    ws.column_dimensions[get_column_letter(src_col + 1)].hidden = True
    ws.column_dimensions[get_column_letter(src_col + 2)].hidden = True

    row = C.section_header(ws, "Usage and credit trend", row=row, last_col=LAST) + 1
    if len(daily) >= 2:
        last_data = 1 + len(daily)
        line = LineChart()
        line.title = "Usage and credit trend"
        line.y_axis.title = "Count"
        line.x_axis.title = "Date"
        _finish_axes(line, x_title="Date", y_title="Count")
        line.height = 8
        line.width = 18
        line.style = 10
        cats = Reference(ws, min_col=src_col, min_row=2, max_row=last_data)
        data = Reference(ws, min_col=src_col + 1, min_row=1, max_col=src_col + 2, max_row=last_data)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.shape = 4
        ws.add_chart(line, f"A{row}")
        row += 16

    tools = snapshot.get("tools") or []
    row = C.section_header(ws, "Tool usage distribution", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Tool", 16, key="tool"),
        C.Col("Category", 14, key="category"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
        C.Col("Users", 10, "right", T.FMT_INT, key="users"),
        C.Col("Adoption %", 12, "right", T.FMT_DECIMAL1, key="adoptionPct"),
        C.Col("Growth %", 12, "right", T.FMT_DECIMAL1, key="growthPct"),
    ], tools[:12], start_row=row, table_name="DashTools")

    teams = snapshot.get("teams") or []
    row = C.section_header(ws, "Team comparison", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Team", 22, key="name"),
        C.Col("Lead", 18, key="lead"),
        C.Col("Users", 10, "right", T.FMT_INT, key="users"),
        C.Col("Active", 10, "right", T.FMT_INT, key="activeUsers"),
        C.Col("Hours", 12, "right", T.FMT_DECIMAL1, key="usageHours"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
    ], teams, start_row=row, table_name="DashTeams")
    C.freeze_below(ws, 4)


def _sheet_summary(ws, snapshot):
    C.hide_gridlines(ws)
    C.set_widths(ws, [28, 22, 18, 18, 18, 18, 18, 18, 18, 18, 18, 18])
    k = snapshot["kpis"]
    period = snapshot["period"]
    row = C.title_band(ws, "Executive Summary", period["label"], last_col=LAST)
    row = C.section_header(ws, "Organisation snapshot", row=row, last_col=LAST)
    rows = [
        {"metric": "Total users in scope", "value": k["totalUsers"]["value"], "previous": k["totalUsers"]["previous"], "change": k["totalUsers"]["deltaPct"]},
        {"metric": "Active users", "value": k["activeUsers"]["value"], "previous": k["activeUsers"]["previous"], "change": k["activeUsers"]["deltaPct"]},
        {"metric": "Teams", "value": k["teams"], "previous": "—", "change": None},
        {"metric": "Usage hours", "value": k["usageHours"]["value"], "previous": k["usageHours"]["previous"], "change": k["usageHours"]["deltaPct"]},
        {"metric": "Generations", "value": k["generations"]["value"], "previous": k["generations"]["previous"], "change": k["generations"]["deltaPct"]},
        {"metric": "Credits consumed", "value": k["credits"]["value"], "previous": k["credits"]["previous"], "change": k["credits"]["deltaPct"]},
        {"metric": "Average hours / user", "value": k["avgHoursPerUser"]["value"], "previous": k["avgHoursPerUser"]["previous"], "change": k["avgHoursPerUser"]["deltaPct"]},
        {"metric": "Average credits / user", "value": k["avgCreditsPerUser"]["value"], "previous": k["avgCreditsPerUser"]["previous"], "change": k["avgCreditsPerUser"]["deltaPct"]},
        {"metric": "Success rate %", "value": k["successRate"]["value"], "previous": k["successRate"]["previous"], "change": k["successRate"]["deltaPct"]},
        {"metric": "Most-used tool", "value": k["mostUsedTool"], "previous": "—", "change": None},
        {"metric": "Fastest-growing tool", "value": k["fastestGrowingTool"], "previous": "—", "change": k.get("fastestGrowingPct")},
        {"metric": "Most expensive tool", "value": k["mostExpensiveTool"], "previous": "—", "change": None},
        {"metric": "Least-used tool", "value": k["leastUsedTool"], "previous": "—", "change": None},
    ]
    row = C.data_table(ws, [
        C.Col("Metric", 28, key="metric"),
        C.Col("Current", 18, key="value"),
        C.Col("Previous period", 18, key="previous"),
        C.Col("Change %", 14, "right", key="change"),
    ], rows, start_row=row, table_name="ExecSummary")

    row = C.section_header(ws, "Key findings", row=row, last_col=LAST)
    for i, finding in enumerate(snapshot.get("findings") or ["None"], start=1):
        row = C.label_value(ws, f"{i}.", finding, row=row, value_last_col=LAST)

    row += 1
    row = C.section_header(ws, "How to read this report", row=row, last_col=LAST)
    notes = snapshot.get("methodology") or {}
    for label, text in (
        ("Engagement score", notes.get("engagement")),
        ("Credits", notes.get("credits")),
        ("Usage time", notes.get("usageTime")),
        ("Team", notes.get("team")),
        ("Privacy", notes.get("category")),
    ):
        if text:
            row = C.label_value(ws, label, text, row=row, value_last_col=LAST)
    C.freeze_below(ws, 5)


def _sheet_teams(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(ws, "Team Overview", f"Department comparison · {period['label']}", last_col=LAST)
    teams = snapshot.get("teams") or []
    row = C.data_table(ws, [
        C.Col("Team", 22, key="name"),
        C.Col("Lead", 18, key="lead"),
        C.Col("Users", 10, "right", T.FMT_INT, key="users"),
        C.Col("Active", 10, "right", T.FMT_INT, key="activeUsers"),
        C.Col("Inactive", 10, "right", T.FMT_INT, key="inactiveUsers"),
        C.Col("Hours", 12, "right", T.FMT_DECIMAL1, key="usageHours"),
        C.Col("Sessions", 12, "right", T.FMT_INT, key="sessions"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Avg credits/user", 16, "right", T.FMT_DECIMAL1, key="avgCreditsPerUser"),
        C.Col("Avg hours/user", 14, "right", T.FMT_DECIMAL1, key="avgHoursPerUser"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
    ], teams, start_row=row, table_name="Teams")
    C.freeze_below(ws, 4)
    if teams:
        _color_scale(ws, 9, 4, 3 + len(teams))


def _sheet_users(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "User Overview",
        f"Sortable employee metrics · {period['label']} · Engagement is not a productivity score",
        last_col=14,
    )
    users = snapshot.get("users") or []
    row = C.data_table(ws, [
        C.Col("Name", 22, key="name"),
        C.Col("Employee ID", 14, key="employeeId"),
        C.Col("Team", 16, key="department"),
        C.Col("Role", 12, key="role"),
        C.Col("Status", 10, key="accountStatus"),
        C.Col("Hours", 10, "right", T.FMT_DECIMAL1, key="usageHours"),
        C.Col("Active days", 12, "right", T.FMT_INT, key="activeDays"),
        C.Col("Generations", 13, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
        C.Col("Credits / success", 16, "right", T.FMT_DECIMAL1, key="creditsPerGeneration"),
        C.Col("Tools", 10, "right", T.FMT_INT, key="toolsUsed"),
        C.Col("Engagement", 12, "right", T.FMT_INT, key="engagementScore"),
        C.Col("Band", 12, key="usageBand"),
    ], users, start_row=row, table_name="Users")
    C.freeze_below(ws, 4)


def _sheet_individual(ws, snapshot):
    C.hide_gridlines(ws)
    C.set_widths(ws, [22, 18, 16, 14, 14, 14, 14, 16, 14, 14, 14, 14])
    period = snapshot["period"]
    focus = snapshot.get("individual")
    users = [focus] if focus else (snapshot.get("users") or [])
    row = C.title_band(
        ws, "Individual User Report",
        f"{'One user' if focus else 'All users in scope'} · {period['label']}",
        last_col=LAST,
    )
    rendered = 0
    for user in users:
        if rendered >= 40:
            row = C.callout(ws, "Additional users are listed on User Overview. Filter the report to a single employee for a full profile.", row=row, last_col=LAST)
            break
        rendered += 1
        row = C.section_header(ws, f"{user['name']}  ·  {user['employeeId']}", row=row, last_col=LAST)
        info = [
            ("Email", user.get("email") or "—"),
            ("Team", user.get("department") or "—"),
            ("Role", user.get("role") or "—"),
            ("Department lead", user.get("teamLead") or "—"),
            ("Account", user.get("accountStatus") or "—"),
            ("Report period", period["label"]),
            ("Active days", user.get("activeDays")),
            ("Sessions", user.get("sessions")),
            ("Usage hours", user.get("usageHours")),
            ("Avg daily hours", user.get("avgDailyHours")),
            ("First activity", user.get("firstActivity") or "—"),
            ("Last activity", user.get("lastActivity") or "—"),
            ("Tools used", user.get("toolsUsed")),
            ("Primary tool", user.get("primaryTool")),
            ("Primary category", user.get("primaryCategory")),
            ("Generations", user.get("generations")),
            ("Successful", user.get("successfulGenerations")),
            ("Failed", user.get("failedGenerations")),
            ("Credits", user.get("credits")),
            ("Credits / success", user.get("creditsPerGeneration") if user.get("creditsPerGeneration") is not None else "—"),
            ("Engagement (not productivity)", user.get("engagementScore")),
        ]
        for label, value in info:
            row = C.label_value(ws, label, value if value is not None else "—", row=row, value_last_col=4)
        row += 1
        tools = user.get("tools") or []
        if tools:
            row = C.data_table(ws, [
                C.Col("Tool", 16, key="tool"),
                C.Col("Category", 14, key="category"),
                C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
                C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
                C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
                C.Col("Last used", 14, key="lastUsed"),
                C.Col("Time on tool", 18, key="timeSpentNote"),
            ], tools, start_row=row, table_name=f"UserTools{user['userId']}")
        row += 1


def _sheet_tools(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(ws, "Tool Usage", f"Adoption, volume and efficiency · {period['label']}", last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Tool", 16, key="tool"),
        C.Col("Category", 14, key="category"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Success", 12, "right", T.FMT_INT, key="success"),
        C.Col("Failed", 10, "right", T.FMT_INT, key="failed"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Users", 10, "right", T.FMT_INT, key="users"),
        C.Col("Adoption %", 12, "right", T.FMT_DECIMAL1, key="adoptionPct"),
        C.Col("Growth %", 12, "right", T.FMT_DECIMAL1, key="growthPct"),
        C.Col("Credits / success", 16, "right", T.FMT_DECIMAL1, key="creditsPerSuccess"),
        C.Col("Last used", 14, key="lastUsed"),
    ], snapshot.get("tools") or [], start_row=row, table_name="Tools")
    C.freeze_below(ws, 4)


def _sheet_consolidated(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Consolidated",
        f"Every tool × employee pairing in scope, with usage and when it was last used · {period['label']}",
        last_col=LAST,
    )
    rows = []
    for user in snapshot.get("users") or []:
        for t in user.get("tools") or []:
            last_used = t.get("lastUsed")
            if not last_used or last_used == "None":
                continue
            rows.append({
                "tool": t.get("tool"),
                "category": _blank(t.get("category"), "Uncategorised"),
                "name": user.get("name"),
                "department": user.get("department"),
                "generations": t.get("generations") or 0,
                "credits": round(t.get("credits") or 0, 2),
                "lastUsed": last_used,
            })
    rows.sort(key=lambda r: (r["tool"] or "", r["name"] or ""))
    row = C.section_header(ws, "Employee tool usage", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Tool", 18, key="tool"),
        C.Col("Category", 16, key="category"),
        C.Col("Employee name", 24, key="name"),
        C.Col("Team", 18, key="department"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
        C.Col("Last used", 14, key="lastUsed"),
    ], rows, start_row=row, table_name="Consolidated")

    teams = snapshot.get("teams") or []
    row = C.section_header(ws, "Department-wise usage", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Team", 22, key="name"),
        C.Col("Users", 10, "right", T.FMT_INT, key="users"),
        C.Col("Active", 10, "right", T.FMT_INT, key="activeUsers"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
        C.Col("Hours", 12, "right", T.FMT_DECIMAL1, key="usageHours"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
    ], teams, start_row=row, table_name="ConsolidatedDepts")

    departments = snapshot.get("departmentUsage") or []
    if departments:
        row = C.section_header(ws, "Department detail — members, tools and clients", row=row, last_col=LAST)
        for i, entry in enumerate(departments):
            dept_rows = entry.get("rows") or []
            if not dept_rows:
                continue
            row = C.section_header(
                ws,
                f"{entry['department']}  ·  {entry.get('totalGenerations', 0)} generations  ·  {entry.get('totalCredits', 0)} credits  ·  ₹{entry.get('totalCostRupees', 0)}",
                row=row, last_col=LAST,
            )
            row = C.data_table(ws, [
                C.Col("Member", 24, key="name"),
                C.Col("Tool", 16, key="tool"),
                C.Col("Client", 22, key="client"),
                C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
                C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
                C.Col("Cost (₹)", 14, "right", T.FMT_DECIMAL1, key="costRupees"),
            ], dept_rows, start_row=row, table_name=f"Dept{i}_{_safe_file_stub(entry['department'])[:18]}")
            row += 1

    C.freeze_below(ws, 4)


def _sheet_credits(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Credit Analysis",
        f"Efficiency = credits per successful generation. Lower is better. · {period['label']}",
        last_col=LAST,
    )
    users = sorted(snapshot.get("users") or [], key=lambda u: u.get("credits") or 0, reverse=True)
    row = C.data_table(ws, [
        C.Col("Name", 22, key="name"),
        C.Col("Team", 16, key="department"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Successful gens", 16, "right", T.FMT_INT, key="successfulGenerations"),
        C.Col("Credits / success", 16, "right", T.FMT_DECIMAL1, key="creditsPerGeneration"),
        C.Col("Outputs / credit", 16, "right", key="outputsPerCredit"),
        C.Col("Credits / active day", 18, "right", T.FMT_DECIMAL1, key="creditsPerActiveDay"),
        C.Col("Band", 12, key="usageBand"),
    ], users, start_row=row, table_name="Credits")
    C.freeze_below(ws, 4)


def _sheet_generations(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Generation Analysis",
        f"What users generated (tool-derived categories; prompts are not included) · {period['label']}",
        last_col=LAST,
    )
    row = C.section_header(ws, "By category", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Category", 22, key="category"),
        C.Col("Generations", 16, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
    ], snapshot.get("categories") or [], start_row=row, table_name="Categories")
    row = C.section_header(ws, "By tool", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Tool", 16, key="tool"),
        C.Col("Category", 14, key="category"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Failed", 10, "right", T.FMT_INT, key="failed"),
        C.Col("Success %", 12, "right", T.FMT_DECIMAL1, key="successRate"),
    ], snapshot.get("tools") or [], start_row=row, table_name="GenByTool")


def _sheet_timeline(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    rows = snapshot.get("timeline") or []
    row = C.title_band(
        ws, "Activity Timeline",
        f"User × tool × day (capped, no prompt text) · {period['label']}",
        last_col=LAST,
    )
    row = C.data_table(ws, [
        C.Col("Date", 14, key="date"),
        C.Col("User", 22, key="userName"),
        C.Col("Team", 16, key="department"),
        C.Col("Tool", 14, key="tool"),
        C.Col("Category", 14, key="category"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits (allocated)", 18, "right", T.FMT_DECIMAL1, key="credits"),
        C.Col("Status", 12, key="status"),
    ], rows, start_row=row, table_name="Timeline")
    C.freeze_below(ws, 4)


def _sheet_trends(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    daily = (snapshot.get("trends") or {}).get("daily") or []
    row = C.title_band(ws, "Trends", f"Daily usage, generations and credits · {period['label']}", last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Date", 14, key="date"),
        C.Col("Active users", 14, "right", T.FMT_INT, key="activeUsers"),
        C.Col("Usage hours", 14, "right", T.FMT_DECIMAL1, key="usageHours"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Credits", 12, "right", T.FMT_DECIMAL1, key="credits"),
    ], daily, start_row=row, table_name="Trends")
    if len(daily) >= 2:
        last = 3 + len(daily)
        chart = LineChart()
        chart.y_axis.title = "Volume"
        chart.x_axis.title = "Date"
        chart.height = 8
        chart.width = 18
        cats = Reference(ws, min_col=1, min_row=4, max_row=last)
        data = Reference(ws, min_col=4, min_row=3, max_col=5, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "G3")
    C.freeze_below(ws, 4)


def _sheet_anomalies(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Anomalies & Alerts",
        f"Flags requiring review — not automatic findings of misuse · {period['label']}",
        last_col=LAST,
    )
    row = C.data_table(ws, [
        C.Col("Severity", 12, key="severity"),
        C.Col("Kind", 16, key="kind"),
        C.Col("User", 22, key="userName"),
        C.Col("Team", 16, key="department"),
        C.Col("Finding", 60, key="finding"),
        C.Col("Recommended action", 50, key="recommendedAction"),
        C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
    ], snapshot.get("anomalies") or [], start_row=row, table_name="Anomalies")
    C.freeze_below(ws, 4)


def _sheet_actions(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Management Actions",
        f"What management should do next · {period['label']}",
        last_col=LAST,
    )
    row = C.data_table(ws, [
        C.Col("Priority", 10, "center", T.FMT_INT, key="priority"),
        C.Col("Kind", 16, key="kind"),
        C.Col("Target", 22, key="target"),
        C.Col("Team", 16, key="department"),
        C.Col("Finding", 50, key="title"),
        C.Col("Recommended action", 55, key="action"),
    ], snapshot.get("actions") or [], start_row=row, table_name="Actions")
    C.freeze_below(ws, 4)


def _sheet_raw(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Raw Data",
        f"Underlying aggregated records used to generate this report. No prompts, tokens or credentials. · {period['label']}",
        last_col=LAST,
    )
    note = snapshot.get("methodology") or {}
    row = C.callout(ws, note.get("generations") or "", row=row, last_col=LAST)
    rows = snapshot.get("timeline") or []
    row = C.data_table(ws, [
        C.Col("Date", 14, key="date"),
        C.Col("User ID", 10, "right", T.FMT_INT, key="userId"),
        C.Col("User", 22, key="userName"),
        C.Col("Team", 16, key="department"),
        C.Col("Tool", 14, key="tool"),
        C.Col("Category", 14, key="category"),
        C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
        C.Col("Allocated credits", 18, "right", T.FMT_DECIMAL1, key="credits"),
    ], rows, start_row=row, table_name="RawData")
    C.freeze_below(ws, 5)


def _delta(metric: dict) -> str:
    if not metric:
        return ""
    pct = metric.get("deltaPct")
    if pct is None:
        return ""
    sign = "+" if pct > 0 else ""
    return f"{sign}{pct}% vs prior period"


def _color_scale(ws, col: int, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    letter = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{letter}{start_row}:{letter}{end_row}",
        ColorScaleRule(
            start_type="min", start_color="FCE4D6",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="C6EFCE",
        ),
    )


# --------------------------------------------------------------------------- #
# Standalone consolidated report (tool x employee x last used, one tab)
# --------------------------------------------------------------------------- #
def _build_consolidated_workbook(snapshot: dict) -> tuple[bytes, str, str]:
    C._used_table_names.clear()
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Consolidated")
    ws.sheet_properties.tabColor = T.GREEN
    _sheet_consolidated(ws, snapshot)

    ws2 = wb.create_sheet(title="Client Usage")
    ws2.sheet_properties.tabColor = T.NAVY_TABLE
    _sheet_client_usage(ws2, snapshot)

    ws3 = wb.create_sheet(title="Tool Assignments")
    ws3.sheet_properties.tabColor = T.GREEN
    _sheet_tool_assignments(ws3, snapshot)

    period = snapshot["period"]
    wb.properties.title = "Consolidated Usage Report"
    wb.properties.creator = "Task Dashboard — Usage Intelligence"
    wb.properties.subject = f"Tool x employee last-used · {period['label']}"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"Consolidated_{period['end']}.xlsx"
    return buf.getvalue(), XLSX_MIMETYPE, filename


def _sheet_client_usage(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Client Usage",
        f"One table per client — who used which tool and how many credits · {period['label']}",
        last_col=LAST,
    )
    clients = snapshot.get("clientUsage") or []
    if not clients:
        C.callout(ws, "No client-linked generations in this period.", row=row, last_col=LAST)
        return
    for i, entry in enumerate(clients):
        rows = entry.get("rows") or []
        if not rows:
            continue
        row = C.section_header(
            ws,
            f"{entry['client']}  ·  {entry.get('totalGenerations', 0)} generations  ·  {entry.get('totalCredits', 0)} credits  ·  ₹{entry.get('totalCostRupees', 0)}",
            row=row, last_col=LAST,
        )
        row = C.data_table(ws, [
            C.Col("User", 24, key="name"),
            C.Col("Tool", 18, key="tool"),
            C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
            C.Col("Credits", 14, "right", T.FMT_INT, key="credits"),
            C.Col("Cost (₹)", 14, "right", T.FMT_DECIMAL1, key="costRupees"),
        ], rows, start_row=row, table_name=f"Client{i}_{_safe_file_stub(entry['client'])[:18]}")
        row += 1


def _sheet_tool_assignments(ws, snapshot):
    C.hide_gridlines(ws)
    period = snapshot["period"]
    row = C.title_band(
        ws, "Tool Assignments",
        f"Every employee assigned a tool and its account, whether used yet or not · {period['label']}",
        last_col=LAST,
    )

    assignments = snapshot.get("toolAssignments") or []
    row = C.section_header(ws, "Tool assignment by employee", row=row, last_col=LAST)
    row = C.data_table(ws, [
        C.Col("Tool", 18, key="tool"),
        C.Col("User", 24, key="name"),
        C.Col("Department", 20, key="department"),
        C.Col("Account", 30, key="account"),
    ], assignments, start_row=row, table_name="ToolAssignments")

    accounts = snapshot.get("toolAccounts") or []
    row = C.section_header(ws, "Tool accounts", row=row, last_col=LAST)
    row = C.callout(
        ws,
        "Status filter defaults to Current only. Use the Status column filter to also show Old (unassigned) accounts.",
        row=row, last_col=LAST,
    )
    accounts_header_row = row
    row = C.data_table(ws, [
        C.Col("Tool", 20, key="tool"),
        C.Col("Account", 30, key="account"),
        C.Col("Status", 14, key="status"),
        C.Col("Renewal date", 16, key="renewalDate"),
    ], accounts, start_row=row, table_name="ToolAccounts")
    _filter_to_current(ws, "ToolAccounts", accounts_header_row, accounts, status_col_index=2)
    C.freeze_below(ws, 4)


def _filter_to_current(ws, table_name: str, header_row: int, rows: list[dict], *, status_col_index: int) -> None:
    """Pre-apply an Excel table filter so only Status == 'Current' rows show on open."""
    if not rows:
        return
    table = ws.tables.get(table_name)
    if table is None:
        return
    table.autoFilter = AutoFilter(ref=table.ref)
    table.autoFilter.filterColumn.append(FilterColumn(colId=status_col_index, filters=Filters(filter=["Current"])))
    for i, r in enumerate(rows):
        if r.get("status") != "Current":
            ws.row_dimensions[header_row + 1 + i].hidden = True


# --------------------------------------------------------------------------- #
# Single-tab individual report (team-lead walkthrough)
# --------------------------------------------------------------------------- #
def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[:\\/?*\[\]]', " ", name or "Person").strip() or "Person"
    return cleaned[:31]


def _safe_file_stub(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name or "Person").strip("_")
    return (cleaned or "Person")[:40]


def _build_person_workbook(snapshot: dict, person: dict) -> tuple[bytes, str, str]:
    C._used_table_names.clear()
    wb = Workbook()
    wb.remove(wb.active)
    title = _safe_sheet_name(person.get("name") or "Individual")
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = T.NAVY
    _render_person_sheet(ws, snapshot, person)

    period = snapshot["period"]
    wb.properties.title = f"{person.get('name') or 'Individual'}: usage report"
    wb.properties.creator = "Task Dashboard Usage Intelligence"
    wb.properties.subject = f"Individual usage · {period['label']}"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"Person-Report_{_safe_file_stub(person.get('name'))}_{period['end']}.xlsx"
    return buf.getvalue(), XLSX_MIMETYPE, filename


def _render_person_sheet(ws, snapshot: dict, person: dict) -> None:
    C.hide_gridlines(ws)
    C.set_widths(ws, [22, 20, 16, 14, 14, 14, 14, 16, 14, 14, 14, 16])
    period = snapshot["period"]
    tools = [
        {
            **t,
            "category": _blank(t.get("category"), "Uncategorised"),
            "lastUsed": _blank(t.get("lastUsed"), "None"),
            "timeSpentNote": _blank(t.get("timeSpentNote"), "0 min"),
        }
        for t in (person.get("tools") or [])
    ]
    daily = _person_daily(snapshot, person, period)

    row = C.title_band(
        ws,
        f"{person.get('name') or 'Employee'}: Individual usage report",
        f"{period['label']}",
        last_col=LAST,
    )

    row = C.section_header(ws, "1. Who this is", row=row, last_col=LAST)
    identity = [
        {"label": "Name", "value": _blank(person.get("name"))},
        {"label": "Employee ID", "value": _blank(person.get("employeeId"))},
        {"label": "Email", "value": _blank(person.get("email"))},
        {"label": "Team", "value": _blank(person.get("department"))},
        {"label": "Role", "value": _blank(person.get("role"))},
        {"label": "Department lead", "value": _blank(person.get("teamLead"))},
        {"label": "Account status", "value": _blank(person.get("accountStatus"))},
        {"label": "Report period", "value": _blank(period.get("label"))},
        {"label": "First activity in period", "value": _blank(person.get("firstActivity"), "None")},
        {"label": "Last activity in period", "value": _blank(person.get("lastActivity"), "None")},
        {"label": "Time using tools", "value": _blank(person.get("toolTimeLabel"), "0 min")},
        {"label": "Top client (by credits)", "value": _blank(person.get("primaryClient"), "Not linked")},
    ]
    row = C.data_table(ws, [
        C.Col("Field", 22, key="label"),
        C.Col("Value", 28, key="value"),
    ], identity, start_row=row, table_name="Who")

    row = C.section_header(ws, "2. Headline numbers", row=row, last_col=LAST)
    row = C.kpi_cards(ws, [
        C.Kpi("Time using tools", person.get("toolTimeHours") or 0, T.FMT_DECIMAL1),
        C.Kpi("Active days", person.get("activeDays") or 0, T.FMT_INT),
        C.Kpi("Generations", person.get("generations") or 0, T.FMT_INT),
        C.Kpi("Credits used", person.get("credits") or 0, T.FMT_INT),
    ], row=row, span=3)
    row = C.kpi_cards(ws, [
        C.Kpi("Successful outputs", person.get("successfulGenerations") or 0, T.FMT_INT),
        C.Kpi("Failed generations", person.get("failedGenerations") or 0, T.FMT_INT),
        C.Kpi("Success rate", (person.get("successRate") or 0) / 100.0, T.FMT_PCT1),
        C.Kpi("Tools used", person.get("toolsUsed") or 0, T.FMT_INT),
    ], row=row, span=3)
    row = C.kpi_cards(ws, [
        C.Kpi("Top client", person.get("primaryClient") or "Not linked"),
        C.Kpi("Credits on top client", person.get("primaryClientCredits") or 0, T.FMT_INT),
        C.Kpi("Not linked generations", person.get("unlinkedGenerations") or 0, T.FMT_INT),
        C.Kpi("Not linked credits", person.get("unlinkedCredits") or 0, T.FMT_INT),
    ], row=row, span=3)

    row = C.section_header(ws, "3. Which tools they used", row=row, last_col=LAST)
    if not tools:
        row = C.callout(ws, "No dashboard launches or AI generations were attributed to this person in the period.", row=row, last_col=LAST)
    else:
        row = C.callout(
            ws,
            "Time is how long they had that tool open after launching it from the dashboard. A session ends at their next launch or after 2 hours.",
            row=row, last_col=LAST,
        )
        tool_header = row
        row = C.data_table(ws, [
            C.Col("Tool", 16, key="tool"),
            C.Col("What they made", 16, key="category"),
            C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
            C.Col("Time (h)", 12, "right", T.FMT_DECIMAL1, key="timeSpentHours"),
            C.Col("Time", 12, key="timeSpentNote"),
            C.Col("Launches", 12, "right", T.FMT_INT, key="launches"),
            C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
            C.Col("Last used", 14, key="lastUsed"),
        ], tools, start_row=row, table_name="PersonTools")
        n = len(tools)
        chart_row = row
        bar = BarChart()
        bar.type = "col"
        bar.title = "Generations by tool"
        bar.style = 10
        bar.height, bar.width = 8, 12
        bar.legend = None
        _finish_axes(bar, x_title="Tool", y_title="Generations")
        data = Reference(ws, min_col=3, min_row=tool_header, max_row=tool_header + n)
        cats = Reference(ws, min_col=1, min_row=tool_header + 1, max_row=tool_header + n)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        if bar.series:
            bar.series[0].graphicalProperties.solidFill = T.NAVY
        ws.add_chart(bar, f"A{chart_row}")
        tbar = BarChart()
        tbar.type = "col"
        tbar.title = "Time using each tool (hours)"
        tbar.style = 10
        tbar.height, tbar.width = 8, 12
        tbar.legend = None
        _finish_axes(tbar, x_title="Tool", y_title="Hours")
        tdata = Reference(ws, min_col=4, min_row=tool_header, max_row=tool_header + n)
        tbar.add_data(tdata, titles_from_data=True)
        tbar.set_categories(cats)
        if tbar.series:
            tbar.series[0].graphicalProperties.solidFill = T.GREEN
        ws.add_chart(tbar, f"G{chart_row}")
        row = chart_row + 16

    clients = person.get("clients") or snapshot.get("clients") or []
    row = C.section_header(ws, "4. Where credits were used (by client)", row=row, last_col=LAST)
    if not clients:
        row = C.callout(ws, "No generations in this period, so there is no client spend to show.", row=row, last_col=LAST)
    else:
        row = C.callout(
            ws,
            "Client is what they picked before generating. Where no client was picked - it was optional for most of this history - the task it was filed under is shown instead. Not linked means neither was set. ChatGPT has no client mapping.",
            row=row, last_col=LAST,
        )
        client_header = row
        row = C.data_table(ws, [
            C.Col("Client", 22, key="client"),
            C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
            C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
            C.Col("Share of credits %", 18, "right", T.FMT_DECIMAL1, key="share"),
            C.Col("Tools used", 22, key="tools"),
        ], clients, start_row=row, table_name="PersonClients")
        n = len(clients)
        chart_row = row
        cbar = BarChart()
        cbar.type = "bar"
        cbar.title = "Credits by client"
        cbar.style = 10
        cbar.height, cbar.width = 8, 15
        cbar.legend = None
        _finish_axes(cbar, x_title="Credits", y_title="Client")
        data = Reference(ws, min_col=3, min_row=client_header, max_row=client_header + n)
        cats = Reference(ws, min_col=1, min_row=client_header + 1, max_row=client_header + n)
        cbar.add_data(data, titles_from_data=True)
        cbar.set_categories(cats)
        if cbar.series:
            cbar.series[0].graphicalProperties.solidFill = T.GREEN
        ws.add_chart(cbar, f"A{chart_row}")
        row = chart_row + 16

    categories = _person_categories(tools)
    row = C.section_header(ws, "5. What kind of work this is", row=row, last_col=LAST)
    if categories:
        cat_header = row
        row = C.data_table(ws, [
            C.Col("Type of output", 22, key="category"),
            C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
            C.Col("Credits", 12, "right", T.FMT_INT, key="credits"),
            C.Col("Share of work %", 16, "right", T.FMT_DECIMAL1, key="share"),
        ], categories, start_row=row, table_name="PersonCats")
        n = len(categories)
        chart_row = row
        pie = PieChart()
        pie.title = "Share of generations by type of output"
        pie.height, pie.width = 8, 12
        pdata = Reference(ws, min_col=2, min_row=cat_header, max_row=cat_header + n)
        pcats = Reference(ws, min_col=1, min_row=cat_header + 1, max_row=cat_header + n)
        pie.add_data(pdata, titles_from_data=True)
        pie.set_categories(pcats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showSerName = False
        pie.dataLabels.showLegendKey = False
        ws.add_chart(pie, f"A{chart_row}")
        row = chart_row + 16
    else:
        row = C.callout(ws, "No categorised output in this period.", row=row, last_col=LAST)

    row = C.section_header(ws, "6. Activity over time", row=row, last_col=LAST)
    if daily:
        day_header = row
        row = C.data_table(ws, [
            C.Col("Date", 14, key="date"),
            C.Col("Generations", 14, "right", T.FMT_INT, key="generations"),
            C.Col("Credits", 12, "right", T.FMT_DECIMAL1, key="credits"),
        ], daily, start_row=row, table_name="PersonDaily")
        n = len(daily)
        chart_row = row
        line = LineChart()
        line.title = "Daily generations and credits"
        line.style = 10
        line.height, line.width = 8, 18
        _finish_axes(line, x_title="Date", y_title="Volume")
        cats = Reference(ws, min_col=1, min_row=day_header + 1, max_row=day_header + n)
        data = Reference(ws, min_col=2, min_row=day_header, max_col=3, max_row=day_header + n)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        ws.add_chart(line, f"A{chart_row}")
        row = chart_row + 18
    else:
        row = C.callout(ws, "No day-by-day generation activity in this period.", row=row, last_col=LAST)

    gen_log = [
        {
            **row,
            "client": _blank(row.get("client"), "Not linked"),
            "category": _blank(row.get("category"), "Uncategorised"),
            "task": _blank(row.get("task"), "None"),
            "status": _blank(row.get("status"), "Unknown"),
            "tool": _blank(row.get("tool"), "Unknown"),
        }
        for row in (person.get("generationLog") or snapshot.get("generationLog") or [])
    ]
    row = C.section_header(ws, "7. Every generation (client + tool)", row=row, last_col=LAST)
    if not gen_log:
        row = C.callout(ws, "No captured generations for this person in the period.", row=row, last_col=LAST)
    else:
        if person.get("generationLogTruncated"):
            row = C.callout(
                ws,
                f"Showing the most recent {person.get('generationLogLimit')} of {person.get('generationLogTotal')} generations. Filter the date range to see the rest.",
                row=row, last_col=LAST,
            )
        row = C.callout(
            ws,
            "One row per captured generation. ChatGPT is one row per conversation (Generations = prompt count) because ChatGPT is not client-mapped. Prompts are not included.",
            row=row, last_col=LAST,
        )
        row = C.data_table(ws, [
            C.Col("Date", 12, key="date"),
            C.Col("Client", 22, key="client"),
            C.Col("Tool", 14, key="tool"),
            C.Col("Type", 14, key="category"),
            C.Col("Generations", 13, "right", T.FMT_INT, key="generations"),
            C.Col("Credits", 12, "right", T.FMT_DECIMAL1, key="credits"),
            C.Col("Status", 14, key="status"),
            C.Col("Task", 22, key="task"),
        ], gen_log, start_row=row, table_name="PersonGens")

    C.set_widths(ws, [22, 22, 16, 16, 16, 16, 14, 16, 14, 14, 14, 16])
    ws.print_title_rows = "1:2"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False
    C.freeze_below(ws, 4)


def _person_daily(snapshot: dict, person: dict, period: dict) -> list[dict]:
    by_day: dict[str, dict] = defaultdict(lambda: {"generations": 0, "credits": 0.0})
    for row in snapshot.get("timeline") or []:
        if row.get("userId") != person.get("userId"):
            continue
        key = str(row.get("date") or "")
        if not key:
            continue
        by_day[key]["generations"] += int(row.get("generations") or 0)
        by_day[key]["credits"] += float(row.get("credits") or 0)
    start = datetime.strptime(period["start"], "%Y-%m-%d").date()
    end = datetime.strptime(period["end"], "%Y-%m-%d").date()
    out = []
    cursor = start
    while cursor <= end:
        key = str(cursor)
        slot = by_day.get(key, {"generations": 0, "credits": 0.0})
        out.append({"date": key, "generations": int(slot["generations"]), "credits": round(float(slot["credits"]), 2)})
        cursor += timedelta(days=1)
    if period.get("days", 0) > 60:
        return [d for d in out if d["generations"] or d["credits"]] or out[-30:]
    return out


def _person_categories(tools: list[dict]) -> list[dict]:
    by_cat: dict[str, dict] = defaultdict(lambda: {"generations": 0, "credits": 0.0})
    for tool in tools:
        name = _blank(tool.get("category"), "Uncategorised")
        by_cat[name]["generations"] += int(tool.get("generations") or 0)
        by_cat[name]["credits"] += float(tool.get("credits") or 0)
    total = sum(v["generations"] for v in by_cat.values()) or 1
    rows = [
        {
            "category": name,
            "generations": vals["generations"],
            "credits": round(vals["credits"], 2),
            "share": round(vals["generations"] / total * 100.0, 1),
        }
        for name, vals in by_cat.items()
        if vals["generations"] or vals["credits"]
    ]
    rows.sort(key=lambda r: r["generations"], reverse=True)
    return rows
